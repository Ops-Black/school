"""
تصدير التقارير إلى Excel
"""
import os
from datetime import datetime

OUT_DIR = os.path.join(os.path.expanduser("~"), "school_pro_reports")
os.makedirs(OUT_DIR, exist_ok=True)


def _wb():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    return Workbook, Font, Alignment, PatternFill, Border, Side


def _header_style(cell, Font, PatternFill, Alignment, color="1D4ED8"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _path(name):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(OUT_DIR, f"{name}_{ts}.xlsx")


def export_students_excel(students):
    Workbook, Font, Alignment, PatternFill, Border, Side = _wb()
    wb = Workbook(); ws = wb.active; ws.title = "الطلاب"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 10

    headers = ["#", "الاسم الكامل", "الصف", "الشعبة", "رقم الهاتف", "ولي الأمر", "الحالة"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        _header_style(cell, Font, PatternFill, Alignment)

    alt = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    for r, s in enumerate(students, 2):
        vals = [r-1, s["full_name"], s["grade"] or "", s["section"] or "",
                s["phone"] or "", s["parent_name"] or "", s["status"] or "نشط"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(horizontal="center")
            if r % 2 == 0: cell.fill = alt

    path = _path("students")
    wb.save(path); return path


def export_fees_excel(fees_data):
    Workbook, Font, Alignment, PatternFill, Border, Side = _wb()
    wb = Workbook(); ws = wb.active; ws.title = "الأقساط"
    for col, w in zip("ABCDEFGH", [5,26,18,8,16,16,16,12]):
        ws.column_dimensions[col].width = w

    headers = ["#","الطالب","الصف","الشعبة","الكلي","المدفوع","المتبقي","الحالة"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        _header_style(cell, Font, PatternFill, Alignment, "166534")

    for r, f in enumerate(fees_data, 2):
        tot = f["total_amount"] or 0
        paid = f["paid_amount"] or 0
        rem = f["remaining"] or 0
        status = "مكتمل" if rem <= 0 else ("جزئي" if paid > 0 else "لم يُسدَّد")
        vals = [r-1, f["full_name"], f["grade"] or "", f["section"] or "",
                int(tot), int(paid), int(rem), status]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(horizontal="center")
            if status == "مكتمل":
                cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            elif status == "لم يُسدَّد":
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    path = _path("fees"); wb.save(path); return path


def export_absence_excel(absence_data):
    Workbook, Font, Alignment, PatternFill, Border, Side = _wb()
    wb = Workbook(); ws = wb.active; ws.title = "الغياب"
    for col, w in zip("ABCD", [26,18,8,14]):
        ws.column_dimensions[col].width = w

    headers = ["الطالب","الصف","الشعبة","عدد الغيابات"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        _header_style(cell, Font, PatternFill, Alignment, "B45309")

    for r, row in enumerate(absence_data, 2):
        vals = [row["full_name"], row["grade"] or "", row["section"] or "", row["absences"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(horizontal="center")
            if row["absences"] >= 10:
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    path = _path("absence"); wb.save(path); return path
