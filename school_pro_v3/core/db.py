"""
قاعدة البيانات الكاملة — متعددة المدارس
"""
import sqlite3, os, json
from datetime import datetime, date

DB_PATH = os.path.join(os.path.expanduser("~"), "school_pro_data.db")

def conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")
    return c

def init():
    db = conn()
    db.executescript("""
    -- ═══ المدارس ═══
    CREATE TABLE IF NOT EXISTS schools (
        id      INTEGER PRIMARY KEY AUTOINCREMENT,
        name    TEXT NOT NULL,
        logo    TEXT,
        address TEXT,
        phone   TEXT,
        email   TEXT,
        principal TEXT,
        academic_year TEXT DEFAULT '2025-2026',
        fee_amount REAL DEFAULT 1500000,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- ═══ المستخدمون ═══
    CREATE TABLE IF NOT EXISTS users (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        school_id INTEGER NOT NULL REFERENCES schools(id),
        username  TEXT NOT NULL,
        password  TEXT NOT NULL,
        full_name TEXT,
        role      TEXT NOT NULL DEFAULT 'موظف',
        perms     TEXT DEFAULT '{}',
        active    INTEGER DEFAULT 1,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(school_id, username)
    );

    -- ═══ الطلاب ═══
    CREATE TABLE IF NOT EXISTS students (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        school_id   INTEGER NOT NULL REFERENCES schools(id),
        full_name   TEXT NOT NULL,
        birth_year  INTEGER,
        gender      TEXT DEFAULT 'ذكر',
        grade       TEXT,
        section     TEXT,
        phone       TEXT,
        address     TEXT,
        parent_name TEXT,
        parent_phone TEXT,
        parent_job  TEXT,
        photo_path  TEXT,
        status      TEXT DEFAULT 'نشط',
        notes       TEXT,
        created_at  TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- ═══ ملفات الطالب ═══
    CREATE TABLE IF NOT EXISTS student_files (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
        file_type  TEXT,   -- 'جنسية', 'بطاقة وطنية', 'شهادة ميلاد', 'صورة', 'أخرى'
        file_path  TEXT,
        uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- ═══ الأقساط ═══
    CREATE TABLE IF NOT EXISTS fees (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id  INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
        total_amount REAL DEFAULT 1500000,
        paid_amount REAL DEFAULT 0,
        UNIQUE(student_id)
    );

    CREATE TABLE IF NOT EXISTS fee_payments (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id  INTEGER NOT NULL REFERENCES students(id),
        amount      REAL NOT NULL,
        payment_date TEXT DEFAULT CURRENT_DATE,
        employee    TEXT,
        notes       TEXT,
        created_at  TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- ═══ الغياب ═══
    CREATE TABLE IF NOT EXISTS attendance (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
        date       TEXT NOT NULL,
        status     TEXT DEFAULT 'حاضر',
        UNIQUE(student_id, date)
    );

    -- ═══ الدرجات ═══
    CREATE TABLE IF NOT EXISTS grades (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER NOT NULL REFERENCES students(id) ON DELETE CASCADE,
        subject    TEXT NOT NULL,
        score      REAL,
        exam_type  TEXT DEFAULT 'فصلي',
        term       TEXT DEFAULT 'الأول',
        year       TEXT,
        UNIQUE(student_id, subject, exam_type, term)
    );

    -- ═══ المدرسون ═══
    CREATE TABLE IF NOT EXISTS teachers (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        school_id      INTEGER NOT NULL REFERENCES schools(id),
        full_name      TEXT NOT NULL,
        specialization TEXT,
        subjects TEXT DEFAULT "[]",  -- JSON list of subjects
        phone          TEXT,
        salary         REAL DEFAULT 0,
        classes_count  INTEGER DEFAULT 0,
        status         TEXT DEFAULT 'نشط',
        created_at     TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- ═══ مواد المدرس ═══
    CREATE TABLE IF NOT EXISTS teacher_subjects (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        teacher_id  INTEGER NOT NULL REFERENCES teachers(id) ON DELETE CASCADE,
        subject     TEXT NOT NULL,
        grade       TEXT,
        section     TEXT
    );

    -- ═══ المحاسبة ═══
    CREATE TABLE IF NOT EXISTS accounting (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        school_id   INTEGER NOT NULL REFERENCES schools(id),
        type        TEXT NOT NULL,   -- 'دخل' أو 'مصروف'
        category    TEXT,            -- 'راتب', 'كهرباء', 'إيجار', ...
        amount      REAL NOT NULL,
        description TEXT,
        date        TEXT DEFAULT CURRENT_DATE,
        created_by  TEXT,
        created_at  TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- ═══ التقويم المدرسي ═══
    CREATE TABLE IF NOT EXISTS calendar_events (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        school_id   INTEGER NOT NULL REFERENCES schools(id),
        title       TEXT NOT NULL,
        event_type  TEXT DEFAULT 'عام',  -- 'امتحان', 'عطلة', 'اجتماع', 'عام'
        start_date  TEXT NOT NULL,
        end_date    TEXT,
        description TEXT,
        color       TEXT DEFAULT '#3B82F6'
    );

    -- ═══ سجل العمليات ═══
    CREATE TABLE IF NOT EXISTS audit_log (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        school_id  INTEGER,
        user_name  TEXT,
        action     TEXT NOT NULL,
        details    TEXT,
        timestamp  TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- ═══ الإشعارات ═══
    CREATE TABLE IF NOT EXISTS notifications (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        school_id  INTEGER NOT NULL,
        type       TEXT,
        message    TEXT,
        student_id INTEGER,
        is_read    INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    """)
    db.commit()
    db.close()

# ─── Audit Log ────────────────────────────────────────────────
def log(school_id, user_name, action, details=""):
    db = conn()
    db.execute("INSERT INTO audit_log (school_id,user_name,action,details) VALUES (?,?,?,?)",
               (school_id, user_name, action, details))
    db.commit(); db.close()

def get_audit_log(school_id, limit=100):
    db = conn()
    rows = db.execute(
        "SELECT * FROM audit_log WHERE school_id=? ORDER BY id DESC LIMIT ?",
        (school_id, limit)).fetchall()
    db.close(); return rows

# ─── Schools ──────────────────────────────────────────────────
def get_schools():
    db = conn()
    rows = db.execute("SELECT * FROM schools ORDER BY id").fetchall()
    db.close(); return rows

def add_school(data):
    db = conn()
    db.execute("""INSERT INTO schools (name,address,phone,email,principal,academic_year,fee_amount)
                  VALUES (:name,:address,:phone,:email,:principal,:academic_year,:fee_amount)""", data)
    sid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    # Default admin for this school
    db.execute("INSERT INTO users (school_id,username,password,full_name,role) VALUES (?,?,?,?,?)",
               (sid, "admin", "admin123", "مدير المدرسة", "مدير"))
    db.commit(); db.close(); return sid

def get_school(school_id):
    db = conn()
    r = db.execute("SELECT * FROM schools WHERE id=?", (school_id,)).fetchone()
    db.close(); return r

def update_school(school_id, data):
    db = conn()
    db.execute("""UPDATE schools SET name=:name, address=:address, phone=:phone,
                  email=:email, principal=:principal, academic_year=:academic_year,
                  fee_amount=:fee_amount WHERE id=:id""", {**data, "id": school_id})
    db.commit(); db.close()

# ─── Users ────────────────────────────────────────────────────
def login(school_id, username, password):
    db = conn()
    r = db.execute("SELECT * FROM users WHERE school_id=? AND username=? AND password=? AND active=1",
                   (school_id, username, password)).fetchone()
    db.close(); return dict(r) if r else None

def get_users(school_id):
    db = conn()
    rows = db.execute("SELECT * FROM users WHERE school_id=? ORDER BY id", (school_id,)).fetchall()
    db.close(); return rows

def add_user(school_id, data):
    db = conn()
    db.execute("INSERT INTO users (school_id,username,password,full_name,role,perms) VALUES (?,?,?,?,?,?)",
               (school_id, data["username"], data["password"], data["full_name"], data["role"], data.get("perms", "{}")))
    db.commit(); db.close()

def delete_user(uid):
    db = conn()
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.commit(); db.close()

# ─── Students ─────────────────────────────────────────────────
def get_students(school_id, search=""):
    db = conn()
    if search:
        q = f"%{search}%"
        rows = db.execute("""SELECT * FROM students WHERE school_id=? AND (
            full_name LIKE ? OR phone LIKE ? OR parent_phone LIKE ? OR
            grade LIKE ? OR section LIKE ? OR address LIKE ?)
            AND status != 'محذوف' ORDER BY full_name""", (school_id,q,q,q,q,q,q)).fetchall()
    else:
        rows = db.execute("SELECT * FROM students WHERE school_id=? AND status!='محذوف' ORDER BY full_name",
                          (school_id,)).fetchall()
    db.close(); return rows

def get_students_with_extras(school_id, search=""):
    """
    نفس get_students لكن يضيف:
      - has_debt: هل عليه قسط متبقٍ
      - absence_count: عدد أيام الغياب الكلي
    يُستخدم في فلترة صفحة الطلاب (عليه قسط / متغيب كثير)
    """
    rows = get_students(school_id, search)
    db = conn()
    out = []
    for r in rows:
        s = dict(r)
        sid = s["id"]
        fee = db.execute("SELECT total_amount, paid_amount FROM fees WHERE student_id=?", (sid,)).fetchone()
        if fee:
            s["has_debt"] = (fee["total_amount"] - (fee["paid_amount"] or 0)) > 0
        else:
            s["has_debt"] = False
        abs_cnt = db.execute(
            "SELECT COUNT(*) c FROM attendance WHERE student_id=? AND status='غائب'", (sid,)
        ).fetchone()
        s["absence_count"] = abs_cnt["c"] if abs_cnt else 0
        out.append(s)
    db.close()
    return out

def get_student(sid):
    db = conn()
    r = db.execute("SELECT * FROM students WHERE id=?", (sid,)).fetchone()
    db.close(); return dict(r) if r else None

def add_student(school_id, data, user_name=""):
    db = conn()
    db.execute("""INSERT INTO students
        (school_id,full_name,birth_year,gender,grade,section,phone,address,parent_name,parent_phone,parent_job,notes)
        VALUES (:school_id,:full_name,:birth_year,:gender,:grade,:section,:phone,:address,:parent_name,:parent_phone,:parent_job,:notes)""",
        {**data, "school_id": school_id})
    sid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    school = db.execute("SELECT fee_amount FROM schools WHERE id=?", (school_id,)).fetchone()
    fee = school["fee_amount"] if school else 1500000
    db.execute("INSERT INTO fees (student_id,total_amount,paid_amount) VALUES (?,?,0)", (sid, fee))
    db.commit(); db.close()
    log(school_id, user_name, "إضافة طالب", data.get("full_name",""))
    return sid

def update_student(sid, data, school_id=0, user_name=""):
    db = conn()
    db.execute("""UPDATE students SET full_name=:full_name, birth_year=:birth_year, gender=:gender,
        grade=:grade, section=:section, phone=:phone, address=:address, parent_name=:parent_name,
        parent_phone=:parent_phone, parent_job=:parent_job, photo_path=:photo_path,
        status=:status, notes=:notes WHERE id=:id""", {**data, "id": sid})
    db.commit(); db.close()
    log(school_id, user_name, "تعديل طالب", data.get("full_name",""))

def delete_student(sid, school_id=0, user_name=""):
    db = conn()
    name = db.execute("SELECT full_name FROM students WHERE id=?", (sid,)).fetchone()
    db.execute("UPDATE students SET status='محذوف' WHERE id=?", (sid,))
    db.commit(); db.close()
    log(school_id, user_name, "حذف طالب", name["full_name"] if name else "")

def import_students_excel(school_id, filepath, user_name=""):
    """
    استيراد ذكي من Excel — يتعرف على أسماء الأعمدة تلقائياً
    يدعم أي ترتيب للأعمدة ويقرأ الدرجات أيضاً
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    count = 0
    errors = 0

    db = conn()
    school = db.execute("SELECT fee_amount FROM schools WHERE id=?", (school_id,)).fetchone()
    fee = school["fee_amount"] if school else 1500000

    # ── Step 1: قراءة صف العناوين وتحديد الأعمدة ──
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        db.close(); return 0

    headers = [str(c).strip() if c is not None else "" for c in header_row]

    # خريطة الأعمدة بالأسماء العربية والإنجليزية
    col = {}
    NAME_KEYS    = {"الاسم الكامل","الاسم","اسم الطالب","name","full_name"}
    GENDER_KEYS  = {"الجنس","gender","sex"}
    GRADE_KEYS   = {"المرحلة","الصف","الصف الدراسي","grade","class","level","المرحلة الدراسية"}
    SECTION_KEYS = {"الشعبة","الفصل","الشعبة الدراسية","section","division"}
    PHONE_KEYS   = {"هاتف الطالب","هاتف","رقم الهاتف","phone","mobile","جوال"}
    PARENT_KEYS  = {"ولي الأمر","اسم ولي الأمر","الأب","parent","parent_name","الوالد"}
    PPHONE_KEYS  = {"هاتف ولي الأمر","هاتف الوالد","parent_phone"}
    ADDR_KEYS    = {"العنوان","المحافظة","address","city","المدينة"}
    BIRTH_KEYS   = {"مواليد","سنة الميلاد","تاريخ الميلاد","birth_year","birth","سنة"}
    JOB_KEYS     = {"مهنة ولي الأمر","المهنة","job","occupation"}

    # مواد الدرجات لاستيرادها تلقائياً
    GRADE_SUBJECTS = {
        "الرياضيات","العربي","اللغة العربية","الإنكليزي","اللغة الإنكليزية",
        "العلوم","الاجتماعيات","التاريخ","الجغرافية","الحاسوب","التربية الإسلامية",
        "الفنية","الرياضة","الفيزياء","الكيمياء","الأحياء","math","arabic","english"
    }
    subject_cols = {}  # subject_name -> col_index

    for i, h in enumerate(headers):
        hl = h.strip()
        if hl in NAME_KEYS:    col["full_name"]    = i
        elif hl in GENDER_KEYS:  col["gender"]       = i
        elif hl in GRADE_KEYS:   col["grade"]        = i
        elif hl in SECTION_KEYS: col["section"]      = i
        elif hl in PHONE_KEYS:   col["phone"]        = i
        elif hl in PARENT_KEYS:  col["parent_name"]  = i
        elif hl in PPHONE_KEYS:  col["parent_phone"] = i
        elif hl in ADDR_KEYS:    col.setdefault("address", i)
        elif hl in BIRTH_KEYS:   col["birth_year"]   = i
        elif hl in JOB_KEYS:     col["parent_job"]   = i
        elif hl in GRADE_SUBJECTS: subject_cols[hl]  = i

    # إذا لم يوجد عمود الاسم، نحاول البحث عن أي عمود نصي في الصف الثاني
    if "full_name" not in col:
        # fallback: العمود الأول غير رقمي بعد الـ ID
        for i, h in enumerate(headers):
            if i == 0: continue  # تخطي ID
            if h and h not in {"ID","id","رقم","#","م"}:
                col["full_name"] = i
                break

    if "full_name" not in col:
        db.close(); return 0

    def cell(row, key, default=""):
        idx = col.get(key)
        if idx is None or idx >= len(row): return default
        v = row[idx]
        return str(v).strip() if v is not None else default

    # ── Step 2: قراءة البيانات ──
    for row in ws.iter_rows(min_row=2, values_only=True):
        # تخطي الصفوف الفارغة
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        name = cell(row, "full_name")
        if not name or name in ("None","nan","الاسم",""):
            continue

        # birth_year
        birth = 2014
        if "birth_year" in col:
            try:
                bv = row[col["birth_year"]]
                if bv:
                    birth = int(str(bv).split("/")[0].split("-")[0])
            except Exception:
                pass

        # استخراج اسم ولي الأمر تلقائياً من اسم الطالب إذا لم يكن موجوداً
        # الطريقة العراقية: "علي عبدالله خالد" → الأب = "عبدالله خالد"
        parent_from_excel = cell(row, "parent_name", "")
        if not parent_from_excel:
            parts = name.strip().split()
            if len(parts) >= 3:
                parent_from_excel = " ".join(parts[1:3])
            elif len(parts) == 2:
                parent_from_excel = parts[1]
            else:
                parent_from_excel = name

        try:
            db.execute("""INSERT INTO students
                (school_id,full_name,birth_year,gender,grade,section,
                 phone,address,parent_name,parent_phone,parent_job,status)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (school_id, name, birth,
                 cell(row,"gender","ذكر"),
                 cell(row,"grade",""),
                 cell(row,"section","أ"),
                 cell(row,"phone",""),
                 cell(row,"address",""),
                 parent_from_excel,
                 cell(row,"parent_phone",""),
                 cell(row,"parent_job",""),
                 "نشط"))
            sid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            db.execute("INSERT OR IGNORE INTO fees (student_id,total_amount,paid_amount) VALUES (?,?,0)",
                       (sid, fee))

            # استيراد الدرجات إذا وُجدت
            for subj, idx in subject_cols.items():
                if idx < len(row) and row[idx] is not None:
                    try:
                        score = float(row[idx])
                        db.execute("""INSERT OR REPLACE INTO grades
                            (student_id,subject,score,exam_type,term,year)
                            VALUES (?,?,?,?,?,?)""",
                            (sid, subj, score, "سنوي", "الأول", "2025-2026"))
                    except Exception:
                        pass

            count += 1
        except Exception as ex:
            errors += 1

    db.commit(); db.close()
    log(school_id, user_name, "استيراد Excel",
        f"{count} طالب" + (f" ({errors} خطأ)" if errors else ""))
    return count

def import_students_word(school_id, filepath, user_name=""):
    """استيراد الطلاب من ملف Word — يقرأ الجداول في الملف"""
    try:
        import docx
    except ImportError:
        return 0, "مكتبة python-docx غير مثبتة"
    doc = docx.Document(filepath)
    count = 0
    db = conn()
    school = db.execute("SELECT fee_amount FROM schools WHERE id=?", (school_id,)).fetchone()
    fee = school["fee_amount"] if school else 1500000
    for table in doc.tables:
        # Try to find header row
        if len(table.rows) < 2:
            continue
        header = [c.text.strip() for c in table.rows[0].cells]
        # Map columns by name
        col_map = {}
        for i, h in enumerate(header):
            if "اسم" in h and "الطالب" in h or h == "الاسم": col_map["full_name"] = i
            elif "مواليد" in h or "الميلاد" in h or "سنة" in h: col_map["birth_year"] = i
            elif "الصف" in h or "صف" in h: col_map["grade"] = i
            elif "الشعبة" in h or "فصل" in h: col_map["section"] = i
            elif "الهاتف" in h or "موبايل" in h: col_map["phone"] = i
            elif "ولي الأمر" in h or "الأب" in h: col_map["parent_name"] = i
            elif "جنس" in h: col_map["gender"] = i
            elif "العنوان" in h: col_map["address"] = i
        if "full_name" not in col_map:
            continue
        for row in table.rows[1:]:
            cells = [c.text.strip() for c in row.cells]
            name = cells[col_map["full_name"]] if col_map.get("full_name") is not None and col_map["full_name"] < len(cells) else ""
            if not name:
                continue
            try:
                birth = int(cells[col_map["birth_year"]]) if col_map.get("birth_year") is not None else 2014
            except Exception:
                birth = 2014
            db.execute("""INSERT INTO students
                (school_id,full_name,birth_year,gender,grade,section,phone,address,parent_name)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (school_id, name, birth,
                 cells[col_map["gender"]] if col_map.get("gender") is not None and col_map["gender"] < len(cells) else "ذكر",
                 cells[col_map["grade"]] if col_map.get("grade") is not None and col_map["grade"] < len(cells) else "",
                 cells[col_map["section"]] if col_map.get("section") is not None and col_map["section"] < len(cells) else "أ",
                 cells[col_map["phone"]] if col_map.get("phone") is not None and col_map["phone"] < len(cells) else "",
                 cells[col_map["address"]] if col_map.get("address") is not None and col_map["address"] < len(cells) else "",
                 cells[col_map["parent_name"]] if col_map.get("parent_name") is not None and col_map["parent_name"] < len(cells) else ""))
            sid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            db.execute("INSERT INTO fees (student_id,total_amount,paid_amount) VALUES (?,?,0)", (sid, fee))
            count += 1
    db.commit(); db.close()
    log(school_id, user_name, "استيراد Word", f"{count} طالب")
    return count, ""

# Student files
def add_student_file(student_id, file_type, file_path):
    db = conn()
    db.execute("INSERT INTO student_files (student_id,file_type,file_path) VALUES (?,?,?)",
               (student_id, file_type, file_path))
    db.commit(); db.close()

def get_student_files(student_id):
    db = conn()
    rows = db.execute("SELECT * FROM student_files WHERE student_id=? ORDER BY id", (student_id,)).fetchall()
    db.close(); return rows

# ─── Fees ─────────────────────────────────────────────────────
def get_fees(school_id):
    db = conn()
    rows = db.execute("""
        SELECT s.id, s.full_name, s.grade, s.section,
               f.total_amount, COALESCE(f.paid_amount,0) as paid_amount,
               (f.total_amount - COALESCE(f.paid_amount,0)) as remaining
        FROM students s LEFT JOIN fees f ON s.id=f.student_id
        WHERE s.school_id=? AND s.status!='محذوف'
        ORDER BY remaining DESC""", (school_id,)).fetchall()
    db.close(); return rows

def add_payment(student_id, amount, employee="", notes="", school_id=0, user_name=""):
    db = conn()
    db.execute("INSERT INTO fee_payments (student_id,amount,employee,notes) VALUES (?,?,?,?)",
               (student_id, amount, employee, notes))
    db.execute("UPDATE fees SET paid_amount=paid_amount+? WHERE student_id=?", (amount, student_id))
    db.commit(); db.close()
    log(school_id, user_name, "تسجيل دفعة", f"طالب #{student_id} — {amount:,.0f} د.ع")
    # Check if still pending → notify
    _check_fee_notification(school_id, student_id)

def _check_fee_notification(school_id, student_id):
    db = conn()
    r = db.execute("SELECT total_amount, paid_amount FROM fees WHERE student_id=?", (student_id,)).fetchone()
    if r and r["paid_amount"] < r["total_amount"]:
        rem = r["total_amount"] - r["paid_amount"]
        s = db.execute("SELECT full_name FROM students WHERE id=?", (student_id,)).fetchone()
        msg = f"الطالب {s['full_name']} — متبقي {int(rem):,} دينار"
        db.execute("INSERT INTO notifications (school_id,type,message,student_id) VALUES (?,?,?,?)",
                   (school_id, "قسط متأخر", msg, student_id))
    db.commit(); db.close()

def get_payments(student_id):
    db = conn()
    rows = db.execute("SELECT * FROM fee_payments WHERE student_id=? ORDER BY id DESC", (student_id,)).fetchall()
    db.close(); return rows

# ─── Attendance ───────────────────────────────────────────────
def get_students_for_attendance(school_id, grade, section):
    db = conn()
    rows = db.execute("""SELECT s.id, s.full_name,
        (SELECT status FROM attendance WHERE student_id=s.id AND date=date('now')) as today_status
        FROM students s WHERE s.school_id=? AND s.grade=? AND s.section=? AND s.status='نشط'
        ORDER BY s.full_name""", (school_id, grade, section)).fetchall()
    db.close(); return rows

def save_attendance(date_str, records, school_id=0, user_name=""):
    db = conn()
    for student_id, status in records.items():
        db.execute("INSERT OR REPLACE INTO attendance (student_id,date,status) VALUES (?,?,?)",
                   (student_id, date_str, status))
        if status == "غائب":
            s = db.execute("SELECT full_name, school_id FROM students WHERE id=?", (student_id,)).fetchone()
            if s:
                # Check total absences → notify if >= 10
                cnt = db.execute("SELECT COUNT(*) FROM attendance WHERE student_id=? AND status='غائب'",
                                 (student_id,)).fetchone()[0]
                if cnt >= 10:
                    db.execute("""INSERT INTO notifications (school_id,type,message,student_id)
                        SELECT ?,?,?,? WHERE NOT EXISTS (
                            SELECT 1 FROM notifications WHERE student_id=? AND type='غيابات متكررة' AND is_read=0)""",
                        (s["school_id"], "غيابات متكررة",
                         f"{s['full_name']} تجاوز 10 غيابات (الإجمالي: {cnt})", student_id, student_id))
    db.commit(); db.close()
    log(school_id, user_name, "تسجيل غياب", date_str)

def get_absence_report(school_id, month=None):
    db = conn()
    if month:
        rows = db.execute("""SELECT s.full_name, s.grade, s.section, COUNT(*) as absences
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE s.school_id=? AND a.status='غائب' AND strftime('%Y-%m',a.date)=?
            GROUP BY a.student_id ORDER BY absences DESC""", (school_id, month)).fetchall()
    else:
        rows = db.execute("""SELECT s.full_name, s.grade, s.section, COUNT(*) as absences
            FROM attendance a JOIN students s ON s.id=a.student_id
            WHERE s.school_id=? AND a.status='غائب'
            GROUP BY a.student_id ORDER BY absences DESC""", (school_id,)).fetchall()
    db.close(); return rows

def get_monthly_absence(school_id):
    db = conn()
    rows = db.execute("""SELECT strftime('%Y-%m', date) as month, COUNT(*) as cnt
        FROM attendance a JOIN students s ON s.id=a.student_id
        WHERE s.school_id=? AND a.status='غائب'
        GROUP BY month ORDER BY month DESC LIMIT 6""", (school_id,)).fetchall()
    db.close(); return rows

# ─── Grades ───────────────────────────────────────────────────
def get_grades(student_id):
    db = conn()
    rows = db.execute("SELECT * FROM grades WHERE student_id=? ORDER BY subject", (student_id,)).fetchall()
    db.close(); return rows

def save_grade(student_id, subject, score, exam_type="فصلي", term="الأول", school_id=0, user_name=""):
    db = conn()
    db.execute("""INSERT OR REPLACE INTO grades (student_id,subject,score,exam_type,term,year)
        VALUES (?,?,?,?,?,?)""", (student_id, subject, score, exam_type, term, "2025-2026"))
    db.commit(); db.close()
    log(school_id, user_name, "إدخال درجة", f"طالب #{student_id} — {subject}: {score}")

def get_pass_rate(school_id):
    db = conn()
    r = db.execute("""
        SELECT
            COUNT(DISTINCT CASE WHEN avg_s.avg >= 50 THEN avg_s.sid END) * 100.0 /
            NULLIF(COUNT(DISTINCT avg_s.sid), 0) as pass_rate
        FROM (
            SELECT g.student_id as sid, AVG(g.score) as avg
            FROM grades g JOIN students s ON s.id=g.student_id
            WHERE s.school_id=? GROUP BY g.student_id
        ) avg_s""", (school_id,)).fetchone()
    db.close()
    return round(r["pass_rate"] or 0, 1)

# ─── Teachers ─────────────────────────────────────────────────
def get_teachers(school_id):
    db = conn()
    rows = db.execute("SELECT * FROM teachers WHERE school_id=? ORDER BY full_name", (school_id,)).fetchall()
    db.close(); return rows

def add_teacher(school_id, data, user_name=""):
    db = conn()
    import json
    subjects = data.get("subjects", [])
    db.execute("""INSERT INTO teachers (school_id,full_name,specialization,subjects,phone,salary,classes_count)
        VALUES (?,?,?,?,?,?,?)""", (school_id, data["full_name"], data.get("specialization",""),
                                  json.dumps(subjects, ensure_ascii=False),
                                  data["phone"], data["salary"], data["classes_count"]))
    tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    # Insert each subject
    for subj_info in subjects:
        if isinstance(subj_info, dict):
            db.execute("INSERT INTO teacher_subjects (teacher_id,subject,grade,section) VALUES (?,?,?,?)",
                       (tid, subj_info.get("subject",""), subj_info.get("grade",""), subj_info.get("section","")))
        else:
            db.execute("INSERT INTO teacher_subjects (teacher_id,subject) VALUES (?,?)", (tid, str(subj_info)))
    db.commit(); db.close()
    log(school_id, user_name, "إضافة مدرس", data["full_name"])

def get_teacher_subjects(teacher_id):
    db = conn()
    rows = db.execute("SELECT * FROM teacher_subjects WHERE teacher_id=? ORDER BY id", (teacher_id,)).fetchall()
    db.close()
    return [dict(r) for r in rows]

def update_teacher(tid, data, school_id=0, user_name=""):
    import json
    subjects = data.get("subjects", [])
    db = conn()
    db.execute("""UPDATE teachers SET full_name=?, specialization=?, subjects=?, phone=?, salary=?, classes_count=?
        WHERE id=?""", (data["full_name"], data.get("specialization",""),
                          json.dumps(subjects, ensure_ascii=False),
                          data["phone"], data["salary"], data["classes_count"], tid))
    db.execute("DELETE FROM teacher_subjects WHERE teacher_id=?", (tid,))
    for subj_info in subjects:
        if isinstance(subj_info, dict):
            db.execute("INSERT INTO teacher_subjects (teacher_id,subject,grade,section) VALUES (?,?,?,?)",
                       (tid, subj_info.get("subject",""), subj_info.get("grade",""), subj_info.get("section","")))
        else:
            db.execute("INSERT INTO teacher_subjects (teacher_id,subject) VALUES (?,?)", (tid, str(subj_info)))
    db.commit(); db.close()
    log(school_id, user_name, "تعديل مدرس", data["full_name"])

def delete_teacher(tid, school_id=0, user_name=""):
    db = conn()
    t = db.execute("SELECT full_name FROM teachers WHERE id=?", (tid,)).fetchone()
    db.execute("DELETE FROM teachers WHERE id=?", (tid,))
    db.commit(); db.close()
    log(school_id, user_name, "حذف مدرس", t["full_name"] if t else "")

# ─── Accounting ───────────────────────────────────────────────
def get_accounting(school_id, year_month=None):
    db = conn()
    if year_month:
        rows = db.execute("""SELECT * FROM accounting WHERE school_id=? AND strftime('%Y-%m',date)=?
            ORDER BY date DESC""", (school_id, year_month)).fetchall()
    else:
        rows = db.execute("SELECT * FROM accounting WHERE school_id=? ORDER BY date DESC LIMIT 200",
                          (school_id,)).fetchall()
    db.close(); return rows

def add_accounting(school_id, data, user_name=""):
    db = conn()
    db.execute("""INSERT INTO accounting (school_id,type,category,amount,description,date,created_by)
        VALUES (?,?,?,?,?,?,?)""", (school_id, data["type"], data["category"],
                                    data["amount"], data["description"], data["date"], user_name))
    db.commit(); db.close()
    log(school_id, user_name, f"محاسبة — {data['type']}", f"{data['category']}: {data['amount']:,.0f}")

def get_accounting_summary(school_id):
    db = conn()
    r = db.execute("""SELECT
        SUM(CASE WHEN type='دخل' THEN amount ELSE 0 END) as income,
        SUM(CASE WHEN type='مصروف' THEN amount ELSE 0 END) as expense
        FROM accounting WHERE school_id=?""", (school_id,)).fetchone()
    db.close()
    income = r["income"] or 0; expense = r["expense"] or 0
    return {"income": income, "expense": expense, "profit": income - expense}

def get_monthly_income(school_id):
    db = conn()
    rows = db.execute("""SELECT strftime('%Y-%m',date) as month,
        SUM(CASE WHEN type='دخل' THEN amount ELSE 0 END) as income,
        SUM(CASE WHEN type='مصروف' THEN amount ELSE 0 END) as expense
        FROM accounting WHERE school_id=?
        GROUP BY month ORDER BY month DESC LIMIT 6""", (school_id,)).fetchall()
    db.close(); return rows

# ─── Calendar ─────────────────────────────────────────────────
def get_events(school_id, month=None):
    db = conn()
    if month:
        rows = db.execute("""SELECT * FROM calendar_events WHERE school_id=? AND
            strftime('%Y-%m',start_date)=? ORDER BY start_date""", (school_id, month)).fetchall()
    else:
        rows = db.execute("SELECT * FROM calendar_events WHERE school_id=? ORDER BY start_date",
                          (school_id,)).fetchall()
    db.close(); return rows

def add_event(school_id, data):
    db = conn()
    db.execute("""INSERT INTO calendar_events (school_id,title,event_type,start_date,end_date,description,color)
        VALUES (?,?,?,?,?,?,?)""", (school_id, data["title"], data["event_type"],
                                    data["start_date"], data.get("end_date",""), data.get("description",""), data.get("color","#3B82F6")))
    db.commit(); db.close()

def delete_event(eid):
    db = conn()
    db.execute("DELETE FROM calendar_events WHERE id=?", (eid,))
    db.commit(); db.close()

# ─── Notifications ────────────────────────────────────────────
def get_notifications(school_id, unread_only=True):
    db = conn()
    q = "SELECT * FROM notifications WHERE school_id=?"
    if unread_only: q += " AND is_read=0"
    q += " ORDER BY id DESC LIMIT 50"
    rows = db.execute(q, (school_id,)).fetchall()
    db.close(); return rows

def mark_read(school_id):
    db = conn()
    db.execute("UPDATE notifications SET is_read=1 WHERE school_id=?", (school_id,))
    db.commit(); db.close()

# ─── Stats ────────────────────────────────────────────────────
def get_stats(school_id):
    db = conn()
    total   = db.execute("SELECT COUNT(*) FROM students WHERE school_id=? AND status='نشط'", (school_id,)).fetchone()[0]
    today   = date.today().isoformat()
    absent  = db.execute("""SELECT COUNT(*) FROM attendance a JOIN students s ON s.id=a.student_id
        WHERE s.school_id=? AND a.date=? AND a.status='غائب'""", (school_id, today)).fetchone()[0]
    fees    = db.execute("""SELECT SUM(f.total_amount) as t, SUM(f.paid_amount) as p
        FROM fees f JOIN students s ON s.id=f.student_id WHERE s.school_id=? AND s.status='نشط'""",
        (school_id,)).fetchone()
    teachers= db.execute("SELECT COUNT(*) FROM teachers WHERE school_id=? AND status='نشط'", (school_id,)).fetchone()[0]
    notifs  = db.execute("SELECT COUNT(*) FROM notifications WHERE school_id=? AND is_read=0", (school_id,)).fetchone()[0]
    db.close()
    total_f = fees["t"] or 0; paid_f = fees["p"] or 0
    return {
        "students": total, "teachers": teachers,
        "absent_today": absent,
        "fees_total": total_f, "fees_paid": paid_f, "fees_pending": total_f - paid_f,
        "notifications": notifs,
        "pass_rate": get_pass_rate(school_id),
    }

def get_grades_by_grade_section(school_id, grade, section):
    db = conn()
    rows = db.execute("""SELECT s.id, s.full_name, AVG(g.score) as avg_score
        FROM students s LEFT JOIN grades g ON s.id=g.student_id
        WHERE s.school_id=? AND s.grade=? AND s.section=?
        GROUP BY s.id ORDER BY avg_score DESC""", (school_id, grade, section)).fetchall()
    db.close(); return rows
